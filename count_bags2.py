import os
os.environ["FLAGS_use_mkldnn"] = "False"

from ultralytics import YOLO
import cv2
import re
from collections import Counter
from paddleocr import PaddleOCR
from openpyxl import Workbook, load_workbook
from datetime import datetime

# ---------------- PATHS ----------------
BASE_PATH = r"D:\Projects\Age group\IFFCO"
MODEL_PATH = os.path.join(BASE_PATH, "best.pt")
VIDEO_PATH = os.path.join(BASE_PATH, "videoplayback.mp4")
EXCEL_PATH = os.path.join(BASE_PATH, "wagon_data.xlsx")

# ---------------- LOAD MODEL ----------------
model = YOLO(MODEL_PATH)
cap = cv2.VideoCapture(VIDEO_PATH)

cv2.namedWindow("IFFCO Cement Monitoring", cv2.WINDOW_NORMAL)

# ---------------- OCR ----------------
ocr = PaddleOCR(use_angle_cls=True, lang='en')

# ---------------- EXCEL SETUP ----------------
if not os.path.exists(EXCEL_PATH):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Date", "Time", "Wagon Number",
        "Door1 Count", "Door2 Count",
        "Total"
    ])
    wb.save(EXCEL_PATH)

# ---------------- COUNTERS ----------------
door1_count = 0
door2_count = 0

counted_d1 = set()
counted_d2 = set()

wagon_number_text = ""
wagon_detected = False
wagon_candidates = []

frame_count = 0

print("Processing...")

while True:
    ret, frame = cap.read()
    if not ret:
        break

    frame_count += 1

    results = model.track(
        frame,
        persist=True,
        tracker="botsort.yaml",
        conf=0.30,
        imgsz=960
    )

    doors = []

    if results[0].boxes.id is not None:

        # -------- COLLECT DOORS --------
        for box, track_id, cls in zip(
            results[0].boxes.xyxy,
            results[0].boxes.id,
            results[0].boxes.cls
        ):
            x1, y1, x2, y2 = map(int, box)
            class_id = int(cls)

            if class_id == 1:  # wagon_door
                doors.append((x1, y1, x2, y2))

        # Sort doors left → right
        doors = sorted(doors, key=lambda d: d[0])

        # -------- DRAW DOORS --------
        for i, (dx1, dy1, dx2, dy2) in enumerate(doors):

            color = (0,255,0) if i == 0 else (0,165,255)

            cv2.rectangle(frame, (dx1, dy1), (dx2, dy2), color, 3)
            cv2.putText(frame, f"D{i+1}",
                        (dx1, dy1 - 10),
                        cv2.FONT_HERSHEY_SIMPLEX,
                        1,
                        (0,255,255), 3)

        # -------- PROCESS OBJECTS --------
        for box, track_id, cls in zip(
            results[0].boxes.xyxy,
            results[0].boxes.id,
            results[0].boxes.cls
        ):

            x1, y1, x2, y2 = map(int, box)
            track_id = int(track_id)
            class_id = int(cls)

            cv2.rectangle(frame, (x1, y1), (x2, y2),
                          (255,255,0), 2)

            # -------- BAG COUNTING ONLY --------
            if class_id == 0:  # cement_bag

                center = ((x1+x2)//2, (y1+y2)//2)

                # Door 1
                if len(doors) > 0:
                    dx1, dy1, dx2, dy2 = doors[0]
                    if dx1 <= center[0] <= dx2 and dy1 <= center[1] <= dy2:
                        if track_id not in counted_d1:
                            counted_d1.add(track_id)
                            door1_count += 1

                # Door 2
                if len(doors) > 1:
                    dx1, dy1, dx2, dy2 = doors[1]
                    if dx1 <= center[0] <= dx2 and dy1 <= center[1] <= dy2:
                        if track_id not in counted_d2:
                            counted_d2.add(track_id)
                            door2_count += 1

            # -------- WAGON OCR --------
            if class_id == 2 and not wagon_detected:

                if frame_count % 5 != 0:
                    continue

                pad = 40
                crop = frame[
                    max(0,y1-pad):min(frame.shape[0],y2+pad),
                    max(0,x1-pad):min(frame.shape[1],x2+pad)
                ]

                crop = cv2.resize(crop, None, fx=2, fy=2)
                gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
                gray = cv2.equalizeHist(gray)

                result = ocr.ocr(gray, cls=True)

                if result and result[0]:
                    for line in result[0]:
                        text = line[1][0].upper()
                        text = re.sub(r'[^A-Z0-9 ]', '', text)

                        tokens = text.split()
                        for token in tokens:
                            if len(token) >= 3:
                                wagon_candidates.append(token)

                if len(wagon_candidates) > 30:
                    counter = Counter(wagon_candidates)
                    stable = [t for t,c in counter.items() if c > 4]
                    if stable:
                        wagon_number_text = " ".join(stable)
                        wagon_detected = True
                        print("Wagon Locked:", wagon_number_text)

    total = door1_count + door2_count

    # -------- DISPLAY --------
    cv2.putText(frame, f"D1: {door1_count}", (40,50),
                cv2.FONT_HERSHEY_SIMPLEX,1,(0,255,0),3)

    cv2.putText(frame, f"D2: {door2_count}", (40,100),
                cv2.FONT_HERSHEY_SIMPLEX,1,(0,165,255),3)

    cv2.putText(frame, f"Total: {total}", (40,150),
                cv2.FONT_HERSHEY_SIMPLEX,1,(0,0,255),3)

    cv2.putText(frame, f"Wagon: {wagon_number_text}", (40,220),
                cv2.FONT_HERSHEY_SIMPLEX,1,(0,0,255),3)

    cv2.imshow("IFFCO Cement Monitoring", frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()

# -------- SAVE FINAL DATA --------
if wagon_number_text != "":
    now = datetime.now()
    total = door1_count + door2_count

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    ws.append([
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S"),
        wagon_number_text,
        door1_count,
        door2_count,
        total
    ])

    wb.save(EXCEL_PATH)
    print("Final Data Saved ✔")

print("\nFinal Summary")
print("D1:", door1_count)
print("D2:", door2_count)
print("Total:", total)
print("Wagon:", wagon_number_text)
