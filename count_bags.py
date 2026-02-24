import os
os.environ["FLAGS_use_mkldnn"] = "False"

from ultralytics import YOLO
import cv2
import re
import numpy as np
from collections import Counter
from paddleocr import PaddleOCR
from openpyxl import Workbook, load_workbook
from datetime import datetime

# ----------------------------
# PATHS
# ----------------------------
BASE_PATH = r"D:\Projects\Age group\IFFCO"
MODEL_PATH = os.path.join(BASE_PATH, "best (3).pt")
VIDEO_PATH = os.path.join(BASE_PATH, "clipped_wagon_video.mp4")
EXCEL_PATH = os.path.join(BASE_PATH, "wagon_data.xlsx")

# ----------------------------
# LOAD MODEL
# ----------------------------
model = YOLO(MODEL_PATH)
cap = cv2.VideoCapture(VIDEO_PATH)

cv2.namedWindow("IFFCO Cement Monitoring", cv2.WINDOW_NORMAL)

# ----------------------------
# OCR
# ----------------------------
ocr = PaddleOCR(use_angle_cls=True, lang='en')

# ----------------------------
# EXCEL SETUP
# ----------------------------
if not os.path.exists(EXCEL_PATH):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Date", "Time", "Wagon Number",
        "Door1 Count", "Door2 Count",
        "NP Bags", "DAP Bags",
        "Total"
    ])
    wb.save(EXCEL_PATH)

# ----------------------------
# COLOR DETECTION
# ----------------------------
def detect_product_color(crop):
    h, w = crop.shape[:2]
    crop = crop[int(h*0.2):int(h*0.8), int(w*0.2):int(w*0.8)]

    hsv = cv2.cvtColor(crop, cv2.COLOR_BGR2HSV)
    saturation_mask = hsv[:, :, 1] > 20

    lower_green = np.array([40, 60, 50])
    upper_green = np.array([85, 255, 255])

    lower_blue = np.array([90, 20, 40])
    upper_blue = np.array([145, 255, 255])

    green_mask = cv2.inRange(hsv, lower_green, upper_green)
    blue_mask = cv2.inRange(hsv, lower_blue, upper_blue)

    green_pixels = cv2.countNonZero(green_mask & saturation_mask)
    blue_pixels = cv2.countNonZero(blue_mask & saturation_mask)

    total_pixels = crop.shape[0] * crop.shape[1]

    if green_pixels / total_pixels > 0.20:
        return "DAP"
    elif blue_pixels / total_pixels > 0.05:
        return "NP"
    else:
        return "Unknown"

# ----------------------------
# COUNTERS & STATE
# ----------------------------
door1_count = 0
door2_count = 0
np_count = 0
dap_count = 0

counted_d1 = set()
counted_d2 = set()

# Persistent tracking dictionary for doors
door_assignments = {}

wagon_number_text = ""
wagon_detected = False
wagon_candidates = []

print("Processing...")

frame_count = 0

while True:
    ret, frame = cap.read()
    if not ret:
        break

    frame_count += 1

    results = model.track(
        frame,
        persist=True,
        tracker="botsort.yaml",
        conf=0.30,
        imgsz=960
    )

    if results[0].boxes.id is not None:

        # ---------------- PASS 1: PERSISTENT DOOR ASSIGNMENT ----------------
        current_frame_unassigned = []
        active_doors_this_frame = {}

        for box, track_id, cls in zip(results[0].boxes.xyxy, results[0].boxes.id, results[0].boxes.cls):
            tid = int(track_id)
            class_id = int(cls)

            if class_id == 1:  # wagon_door
                x1, y1, x2, y2 = map(int, box)
                
                # If we haven't seen this Track ID before
                if tid not in door_assignments:
                    current_frame_unassigned.append({'tid': tid, 'x1': x1, 'box': (x1, y1, x2, y2)})
                else:
                    # If we already named it, use the stored label
                    active_doors_this_frame[tid] = {'label': door_assignments[tid], 'box': (x1, y1, x2, y2)}

        # Sort any brand new doors left-to-right and assign them D1/D2 permanently
        if current_frame_unassigned:
            current_frame_unassigned.sort(key=lambda d: d['x1'])
            for d in current_frame_unassigned:
                next_label = f"D{len(door_assignments) + 1}"
                door_assignments[d['tid']] = next_label
                active_doors_this_frame[d['tid']] = {'label': next_label, 'box': d['box']}

        # Draw labels for the active doors
        for tid, data in active_doors_this_frame.items():
            dx1, dy1, dx2, dy2 = data['box']
            label = data['label']
            color = (0, 255, 0) if label == "D1" else (0, 165, 255)
            
            cv2.rectangle(frame, (dx1, dy1), (dx2, dy2), color, 3)
            cv2.putText(frame, label, (dx1, dy1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 255), 3)


        # ---------------- PASS 2: PROCESS BAGS & WAGON OCR ----------------
        for box, track_id, cls in zip(results[0].boxes.xyxy, results[0].boxes.id, results[0].boxes.cls):
            x1, y1, x2, y2 = map(int, box)
            tid = int(track_id)
            class_id = int(cls)

            if class_id != 1: # Draw box for everything except doors (already drawn)
                cv2.rectangle(frame, (x1, y1), (x2, y2), (255,255,0), 2)

            # ---------------- BAG LOGIC ----------------
            if class_id == 0:
                center = ((x1 + x2)//2, (y1 + y2)//2)
                bag_crop = frame[y1:y2, x1:x2]
                
                if bag_crop.size > 0:
                    product_type = detect_product_color(bag_crop)
                else:
                    product_type = "Unknown"

                # Check bag against persistently active doors
                for door_tid, d_data in active_doors_this_frame.items():
                    dx1, dy1, dx2, dy2 = d_data['box']
                    if dx1 <= center[0] <= dx2 and dy1 <= center[1] <= dy2:
                        label = d_data['label']
                        if label == "D1" and tid not in counted_d1:
                            counted_d1.add(tid)
                            door1_count += 1
                            if product_type == "NP": np_count += 1
                            elif product_type == "DAP": dap_count += 1
                        elif label == "D2" and tid not in counted_d2:
                            counted_d2.add(tid)
                            door2_count += 1
                            if product_type == "NP": np_count += 1
                            elif product_type == "DAP": dap_count += 1

            # ---------------- ROBUST WAGON OCR (DUAL-PIPELINE) ----------------
            if class_id == 2 and not wagon_detected:

                if frame_count % 5 != 0:
                    continue

                pad = 70
                y_min, y_max = max(0, y1-pad), min(frame.shape[0], y2+pad)
                x_min, x_max = max(0, x1-pad), min(frame.shape[1], x2+pad)
                crop = frame[y_min:y_max, x_min:x_max]

                if crop.size == 0: continue

                # --- PIPELINE A: DETAIL (Best for Clear/Night Text) ---
                img_A = cv2.resize(crop, None, fx=1.5, fy=1.5, interpolation=cv2.INTER_CUBIC)
                gray_A = cv2.cvtColor(img_A, cv2.COLOR_BGR2GRAY)
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                input_A = clahe.apply(gray_A)

                # --- PIPELINE B: MORPHOLOGY (Best for Hazy Daytime Text) ---
                img_B = cv2.resize(crop, None, fx=0.9, fy=0.9, interpolation=cv2.INTER_AREA)
                gray_B = cv2.cvtColor(img_B, cv2.COLOR_BGR2GRAY)
                blurred_B = cv2.GaussianBlur(gray_B, (3,3), 0)
                _, binary = cv2.threshold(blurred_B, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
                
                # Dilate to connect broken letters in daylight haze
                kernel = np.ones((2,2), np.uint8)
                input_B = cv2.dilate(binary, kernel, iterations=1)

                attempts = [("Detail", input_A), ("Morphology", input_B)]
                
                best_conf = 0.0
                best_tokens = []

                for name, img_input in attempts:
                    result = ocr.ocr(img_input, cls=True)
                    
                    current_conf = 0.0
                    current_tokens = []
                    count = 0

                    if result and result[0]:
                        for line in result[0]:
                            text = line[1][0].upper()
                            conf = line[1][1]
                            
                            text = re.sub(r'[^A-Z0-9 ]', '', text)
                            tokens = text.split()
                            
                            for token in tokens:
                                if len(token) >= 3:
                                    is_valid = False
                                    if any(c.isdigit() for c in token): 
                                        is_valid = True
                                    elif token in ["BCNA", "BOXN", "BHL", "HSM"]: 
                                        is_valid = True
                                    
                                    if is_valid:
                                        current_tokens.append(token)
                                        current_conf += conf
                                        count += 1
                    
                    if count > 0:
                        avg_conf = current_conf / count
                        if avg_conf > best_conf:
                            best_conf = avg_conf
                            best_tokens = current_tokens
                
                if best_tokens:
                    for token in best_tokens:
                        wagon_candidates.append(token)

                # --- FAST STABILIZATION WITH FALLBACK ---
                buffer_size = 60
                if len(wagon_candidates) > buffer_size:
                    recent = wagon_candidates[-buffer_size:]
                    counter = Counter(recent)
                    
                    # Lowered threshold to 3 to lock faster on hazy text
                    stable_tokens = [token for token, count in counter.items() if count >= 3]

                    codes = [t for t in stable_tokens if any(c.isalpha() for c in t)]
                    nums = [t for t in stable_tokens if not any(c.isalpha() for c in t)]

                    standard_lock = len(stable_tokens) >= 2
                    # Fallback if taking too long and only 1 number is found
                    fallback_lock = (len(nums) >= 1) and (len(wagon_candidates) > buffer_size + 10)

                    if standard_lock or fallback_lock:
                        # 1. Base Temporal Sort
                        stable_tokens.sort(key=lambda x: len(recent) - recent[::-1].index(x))
                        
                        codes = [t for t in stable_tokens if any(c.isalpha() for c in t)]
                        nums = [t for t in stable_tokens if not any(c.isalpha() for c in t)]
                        
                        # ---------------------------------------------------------
                        # 2. SMART INDIAN RAILWAYS SORT (Fixes Jumbled Lines)
                        # ---------------------------------------------------------
                        # Force Primary Wagon Types to always be the FIRST code
                        primary_codes = ["BCNA", "BOXN", "BHL"]
                        codes.sort(key=lambda x: 0 if any(p in x for p in primary_codes) else 1)
                        
                        # Force the longer serial number (6-digits) to always be the FIRST number
                        nums.sort(key=lambda x: len(x), reverse=True)
                        # ---------------------------------------------------------
                        
                        # 3. Zipper Merge
                        formatted_parts = []
                        for i in range(max(len(codes), len(nums))):
                            if i < len(codes): formatted_parts.append(codes[i])
                            if i < len(nums): formatted_parts.append(nums[i])
                                
                        wagon_number_text = " ".join(formatted_parts)
                        wagon_detected = True
                        lock_type = "Fallback" if fallback_lock and not standard_lock else "Standard"
                        print(f"Wagon Locked: {wagon_number_text} (Conf: {best_conf:.2f} | Mode: {lock_type})")
                        
    total = door1_count + door2_count

    # ---------------- DISPLAY ----------------
    cv2.putText(frame, f"D1: {door1_count}", (40,50), cv2.FONT_HERSHEY_SIMPLEX,1,(0,255,0),3)
    cv2.putText(frame, f"D2: {door2_count}", (40,100), cv2.FONT_HERSHEY_SIMPLEX,1,(0,165,255),3)
    cv2.putText(frame, f"Total: {total}", (40,150), cv2.FONT_HERSHEY_SIMPLEX,1,(0,0,255),3)
    cv2.putText(frame, f"NP: {np_count}", (40,200), cv2.FONT_HERSHEY_SIMPLEX,1,(0,255,255),3)
    cv2.putText(frame, f"DAP: {dap_count}", (40,250), cv2.FONT_HERSHEY_SIMPLEX,1,(0,255,0),3)
    cv2.putText(frame, f"Wagon: {wagon_number_text}", (40,300), cv2.FONT_HERSHEY_SIMPLEX,1,(0,0,255),3)

    cv2.imshow("IFFCO Cement Monitoring", frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()

# ---------------- SAVE FINAL DATA ----------------
if wagon_number_text != "":
    now = datetime.now()
    total = door1_count + door2_count

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    ws.append([
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S"),
        wagon_number_text,
        door1_count,
        door2_count,
        np_count,
        dap_count,
        total
    ])

    wb.save(EXCEL_PATH)
    print("Final Data Saved to Excel ✔")

print("Final Summary")
print("D1:", door1_count)
print("D2:", door2_count)
print("NP:", np_count)
print("DAP:", dap_count)
print("Total:", total)
print("Wagon:", wagon_number_text)